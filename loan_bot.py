#!/usr/bin/env python3
"""
Telegram Loan Bot for Educational Institutions
Deploy-ready version with environment variable support
"""

import logging
import json
import re
import os
from datetime import datetime
from enum import Enum
from typing import Dict, List, Optional
import openpyxl
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ConversationHandler, ContextTypes, filters
from telegram.constants import ParseMode

# Load environment variables
load_dotenv()

# Enable logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Configuration from environment variables
BOT_TOKEN = os.getenv('BOT_TOKEN')
DATABASE_PATH = os.getenv('DATABASE_PATH', '/mnt/user-data/uploads/Loan_Bot_Database_NEW.xlsx')
ENVIRONMENT = os.getenv('ENVIRONMENT', 'production')

# Validate bot token
if not BOT_TOKEN:
    logger.error("❌ BOT_TOKEN not found! Set it as environment variable")
    raise ValueError("BOT_TOKEN is required")

logger.info(f"✅ Bot Token: {BOT_TOKEN[:10]}...")
logger.info(f"✅ Environment: {ENVIRONMENT}")

# Conversation states
class State(Enum):
    START = 0
    SELECT_COUNTRY = 1
    SELECT_MFI = 2
    SELECT_LOAN_TYPE = 3
    VIEW_LOAN_INFO = 4
    LOAN_CALCULATOR = 5
    REGISTRATION = 6
    REG_NAME = 7
    REG_SCHOOL = 8
    REG_COUNTRY = 9
    REG_LOCATION = 10
    REG_PHONE = 11
    REG_EMAIL = 12
    REG_LOAN_TYPE = 13
    REG_AMOUNT = 14
    REG_PERIOD = 15
    REG_PRIORITY = 16
    REG_COMMENTS = 17

# Load data from Excel using openpyxl
def load_excel_data(filepath):
    """Load data from Excel file"""
    logger.info(f"Loading data from: {filepath}")
    
    data = {}
    wb = openpyxl.load_workbook(filepath)
    
    # Load MFI Master Data
    ws = wb['MFI Master Data']
    data['MFI Master Data'] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # If first column has data
            data['MFI Master Data'].append({
                'MFI ID': row[0],
                'MFI Name': row[1],
                'Country': row[2],
                'Description': row[3],
                'Phone Number': row[4]
            })
    
    # Load Loan Types
    ws = wb['Loan Types']
    data['Loan Types'] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data['Loan Types'].append({
                'Loan Type ID': row[0],
                'Loan Type Name': row[1],
                'Description': row[2]
            })
    
    # Load Interest Rates & Terms (first row)
    ws = wb['Interest Rates & Terms']
    row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    data['Rates'] = {
        'Annual Interest Rate (%)': row[1],
        'Min Amount (USD)': row[2],
        'Max Amount (USD)': row[3],
        'Repayment Period (Months)': row[4],
        'Processing Fee (%)': row[5]
    }
    
    # Load Requirements
    ws = wb['Requirements']
    data['Requirements'] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data['Requirements'].append({
                'Requirement': row[0],
                'Description': row[1]
            })
    
    # Load Countries
    ws = wb['Countries']
    data['Countries'] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data['Countries'].append({
                'Country Code': row[0],
                'Country Name': row[1],
                'Currency': row[2]
            })
    
    logger.info("✅ All data loaded successfully!")
    return data

# Load Excel data
EXCEL_DATA = load_excel_data(DATABASE_PATH)

# Build lookups
COUNTRIES = {item['Country Code']: item['Country Name'] for item in EXCEL_DATA['Countries']}
MFI_DATA = {item['MFI ID']: item for item in EXCEL_DATA['MFI Master Data']}
MFIS_BY_COUNTRY = {}
for mfi in EXCEL_DATA['MFI Master Data']:
    country = mfi['Country']
    if country not in MFIS_BY_COUNTRY:
        MFIS_BY_COUNTRY[country] = []
    MFIS_BY_COUNTRY[country].append(mfi)

LOAN_TYPES = {item['Loan Type ID']: item for item in EXCEL_DATA['Loan Types']}
LOAN_TYPES_LIST = EXCEL_DATA['Loan Types']
REQUIREMENTS = EXCEL_DATA['Requirements']
RATES = EXCEL_DATA['Rates']

def format_currency(amount, currency='USD'):
    """Format currency"""
    symbols = {'USD': '$', 'ETB': 'Br', 'KES': 'Ksh', 'RWF': 'FRw', 'UGX': 'Ush'}
    symbol = symbols.get(currency, currency)
    return f"{symbol}{amount:,.0f}"

def calculate_monthly_payment(principal, annual_rate, months):
    """Calculate monthly payment"""
    monthly_rate = annual_rate / 100 / 12
    if monthly_rate == 0:
        return principal / months
    payment = principal * (monthly_rate * (1 + monthly_rate)**months) / ((1 + monthly_rate)**months - 1)
    return payment

# Telegram Bot Handlers
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Start the conversation"""
    logger.info(f"User {update.effective_user.id} started bot")
    
    keyboard = [
        [InlineKeyboardButton("🏦 MFI List", callback_data='menu_mfis')],
        [InlineKeyboardButton("📊 Loan Calculator", callback_data='menu_calculator')],
        [InlineKeyboardButton("📋 Register", callback_data='menu_register')],
        [InlineKeyboardButton("❓ FAQ", callback_data='menu_faq')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "🤖 *Welcome to Loan Bot* 🤖\n\n"
        "I help schools find microfinance institutions and calculate loans.\n\n"
        "What would you like to do?",
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN
    )
    
    return State.START.value

async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle main menu"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'menu_mfis':
        keyboard = [[InlineKeyboardButton(country, callback_data=f'country_{country}')] 
                   for country in MFIS_BY_COUNTRY.keys()]
        keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data='back_menu')])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "*Select Your Country* 🌍",
            reply_markup=reply_markup,
            parse_mode=ParseMode.MARKDOWN
        )
        return State.SELECT_COUNTRY.value
    
    elif query.data == 'menu_calculator':
        keyboard = [
            [InlineKeyboardButton("EdTech Loan", callback_data='calc_LOAN001')],
            [InlineKeyboardButton("School Development Loan", callback_data='calc_LOAN002')],
            [InlineKeyboardButton("Development Loan", callback_data='calc_LOAN003')],
            [InlineKeyboardButton("⬅️ Back", callback_data='back_menu')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "*Loan Calculator* 📊",
            reply_markup=reply_markup,
            parse_mode=ParseMode.MARKDOWN
        )
        return State.LOAN_CALCULATOR.value
    
    elif query.data == 'menu_register':
        await query.edit_message_text(
            "📋 Registration coming soon!"
        )
        return State.START.value
    
    elif query.data == 'menu_faq':
        await query.edit_message_text(
            "*❓ FAQ*\n\n"
            "*What is this bot?*\n"
            "Helps schools access educational microfinance.\n\n"
            "*Who can apply?*\n"
            "Registered schools with 1+ year operation.\n\n",
            parse_mode=ParseMode.MARKDOWN
        )
        keyboard = [[InlineKeyboardButton("⬅️ Back", callback_data='back_menu')]]
        await query.edit_message_reply_markup(InlineKeyboardMarkup(keyboard))
        return State.START.value
    
    elif query.data == 'back_menu':
        await start(update, context)
        return State.START.value

async def select_country(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle country selection"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'back_menu':
        await start(update, context)
        return State.START.value
    
    country = query.data.replace('country_', '')
    context.user_data['country'] = country
    
    mfis = MFIS_BY_COUNTRY.get(country, [])
    
    message = f"*🏦 MFIs in {country}*\n\n"
    keyboard = []
    
    for mfi in mfis:
        message += f"• *{mfi['MFI Name']}*\n  {mfi['Description']}\n  📞 {mfi['Phone Number']}\n\n"
        keyboard.append([InlineKeyboardButton(mfi['MFI Name'], callback_data=f"mfi_{mfi['MFI ID']}")])
    
    keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data='back_country')])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        message,
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN
    )
    
    return State.SELECT_MFI.value

async def select_mfi(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle MFI selection"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'back_country':
        keyboard = [[InlineKeyboardButton(country, callback_data=f'country_{country}')] 
                   for country in MFIS_BY_COUNTRY.keys()]
        keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data='back_menu')])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            "*Select Your Country* 🌍",
            reply_markup=reply_markup,
            parse_mode=ParseMode.MARKDOWN
        )
        return State.SELECT_COUNTRY.value
    
    mfi_id = query.data.replace('mfi_', '')
    mfi = MFI_DATA[mfi_id]
    context.user_data['mfi_id'] = mfi_id
    context.user_data['mfi_name'] = mfi['MFI Name']
    
    message = f"*{mfi['MFI Name']}*\n\n" \
              f"Contact: {mfi['Phone Number']}\n" \
              f"Country: {mfi['Country']}\n\n" \
              f"*Available Loan Products:*\n\n"
    
    keyboard = []
    for loan in LOAN_TYPES_LIST:
        message += f"💰 *{loan['Loan Type Name']}*\n{loan['Description']}\n\n"
        keyboard.append([InlineKeyboardButton(
            loan['Loan Type Name'],
            callback_data=f"loan_{loan['Loan Type ID']}"
        )])
    
    keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data='back_mfi')])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        message,
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN
    )
    
    return State.SELECT_LOAN_TYPE.value

async def select_loan_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Show loan information"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'back_mfi':
        country = context.user_data.get('country')
        mfis = MFIS_BY_COUNTRY.get(country, [])
        
        message = f"*🏦 MFIs in {country}*\n\n"
        keyboard = []
        
        for mfi in mfis:
            message += f"• *{mfi['MFI Name']}*\n  {mfi['Description']}\n  📞 {mfi['Phone Number']}\n\n"
            keyboard.append([InlineKeyboardButton(mfi['MFI Name'], callback_data=f"mfi_{mfi['MFI ID']}")])
        
        keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data='back_country')])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            message,
            reply_markup=reply_markup,
            parse_mode=ParseMode.MARKDOWN
        )
        return State.SELECT_MFI.value
    
    loan_id = query.data.replace('loan_', '')
    loan = LOAN_TYPES[loan_id]
    mfi_name = context.user_data.get('mfi_name', 'MFI')
    context.user_data['loan_type_id'] = loan_id
    
    message = f"*{loan['Loan Type Name']} @ {mfi_name}*\n\n" \
              f"{loan['Description']}\n\n" \
              f"*📊 Loan Terms:*\n" \
              f"• Annual Interest Rate: *{RATES['Annual Interest Rate (%)']}%*\n" \
              f"• Min Amount: {format_currency(RATES['Min Amount (USD)'])}\n" \
              f"• Max Amount: {format_currency(RATES['Max Amount (USD)'])}\n" \
              f"• Repayment Period: *{RATES['Repayment Period (Months)']} months*\n" \
              f"• Processing Fee: *{RATES['Processing Fee (%)']}%*\n\n" \
              f"*✅ Eligibility Requirements:*\n"
    
    for req in REQUIREMENTS:
        message += f"• {req['Requirement']}\n"
    
    keyboard = [
        [InlineKeyboardButton("⬅️ Back", callback_data='back_loan')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        message,
        reply_markup=reply_markup,
        parse_mode=ParseMode.MARKDOWN
    )
    
    return State.VIEW_LOAN_INFO.value

async def loan_calculator(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle loan calculator"""
    query = update.callback_query
    await query.answer()
    
    if query.data == 'back_loan' or query.data == 'back_menu':
        await start(update, context)
        return State.START.value
    
    loan_id = query.data.replace('calc_', '')
    loan = LOAN_TYPES.get(loan_id, {})
    context.user_data['calc_loan_id'] = loan_id
    
    await query.edit_message_text(
        f"💰 Enter loan amount (between {format_currency(RATES['Min Amount (USD)'])} "
        f"and {format_currency(RATES['Max Amount (USD)'])})"
    )
    
    return State.LOAN_CALCULATOR.value

async def calculator_amount_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Process loan amount"""
    try:
        amount = float(update.message.text.replace(',', ''))
        
        if amount < RATES['Min Amount (USD)'] or amount > RATES['Max Amount (USD)']:
            await update.message.reply_text(f"Amount must be between {format_currency(RATES['Min Amount (USD)'])} and {format_currency(RATES['Max Amount (USD)'])}")
            return State.LOAN_CALCULATOR.value
        
        monthly_payment = calculate_monthly_payment(amount, RATES['Annual Interest Rate (%)'], RATES['Repayment Period (Months)'])
        total_paid = monthly_payment * RATES['Repayment Period (Months)']
        total_interest = total_paid - amount
        
        message = f"*💰 Calculation Results*\n\n" \
                  f"Loan Amount: {format_currency(amount)}\n" \
                  f"Monthly Payment: *{format_currency(monthly_payment)}*\n" \
                  f"Total Interest: {format_currency(total_interest)}\n" \
                  f"Total Amount to Pay: {format_currency(total_paid)}\n"
        
        keyboard = [
            [InlineKeyboardButton("⬅️ Back to Menu", callback_data='back_menu')],
        ]
        
        await update.message.reply_text(
            message,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN
        )
        
        return State.LOAN_CALCULATOR.value
    
    except ValueError:
        await update.message.reply_text("Please enter a valid number")
        return State.LOAN_CALCULATOR.value

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Cancel"""
    await update.message.reply_text("Cancelled.")
    return ConversationHandler.END

def main():
    """Start the bot"""
    logger.info("🚀 Starting Loan Bot...")
    
    if not BOT_TOKEN:
        logger.error("❌ BOT_TOKEN is required!")
        raise ValueError("BOT_TOKEN not set")
    
    application = Application.builder().token(BOT_TOKEN).build()
    logger.info("✅ Application created")
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            State.START.value: [CallbackQueryHandler(main_menu)],
            State.SELECT_COUNTRY.value: [CallbackQueryHandler(select_country)],
            State.SELECT_MFI.value: [CallbackQueryHandler(select_mfi)],
            State.SELECT_LOAN_TYPE.value: [CallbackQueryHandler(select_loan_type)],
            State.VIEW_LOAN_INFO.value: [CallbackQueryHandler(loan_calculator)],
            State.LOAN_CALCULATOR.value: [
                CallbackQueryHandler(loan_calculator),
                MessageHandler(filters.TEXT & ~filters.COMMAND, calculator_amount_input)
            ],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    application.add_handler(conv_handler)
    
    logger.info("✅ Bot configured successfully")
    logger.info("🚀 Bot is now running...")
    
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
